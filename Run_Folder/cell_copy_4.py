# importing openpyxl module
import openpyxl as op;
import os
from pathlib import Path, PureWindowsPath

'''
This script works. It takes the individual sheets and puts the values at the end of the master_test sheet
values.

TO DO:
 - Create for loop so that the program iterates through all of the files in the folder, combining them into the
 master_test sheet
 '''

source_directory = 'C:/Users/sebas/OneDrive/Software/Python/Excel_Auto_3/Correctly_Arranged_Sheets'

# opening the destination excel file
filename1 = "C:\\Users\\sebas\\OneDrive\\Software\\Python\\Excel_Auto_3\\Run_Folder\\master_test.xlsx"
wb2 = op.load_workbook(filename1)
ws2 = wb2.active

for filename in os.listdir(source_directory):
    #name = Path(os.path.abspath(filename)) # gets the abs path for the source files
    name = Path(filename) # gets the abs path for the source files
    print(name)
    '''
    16/9/20
    This isn't working as it doesn'y recognise the file name, very bizarre
    '''

    wb1 = op.load_workbook(name) # loads workbook for the file
    ws1 = wb1.worksheets[0]

    # Calculate # of rows and columns in the source sheet
    # Put this in the for loop so that it updates each iteration
    mr = ws1.max_row
    mc = ws1.max_column

    # Calculate # of rows and columns in the destination sheet
    # Put this in the for loop so that it updates each iteration
    mr_dest = ws2.max_row
    mc_dest = ws2.max_column

	# copying the cell values from source
    # excel file to destination excel file
    for i in range (1, mr + 1):
        for j in range (1, mc + 1):
            # reading cell value from source excel file
            c = ws1.cell(row = i, column = j)
            print(c.value)
            # writing the read value to destination excel file, with
            # the change to have the values get added to the end of the
            # data already in the sheet
            ws2.cell(row = mr_dest + i, column = j).value = c.value

    # saving the destination excel file
    wb2.save(str(filename1))



'''
Has been included in the for loop above
# opening the source excel file
filename ="C:\\Users\\sebas\\OneDrive\\Software\\Python\\Excel_Auto_3\\Correctly_Arranged_Sheets\\Bills_NZHL.xlsx"
wb1 = op.load_workbook(filename)
ws1 = wb1.worksheets[0]
print(filename)
'''
'''
Has been included in the for loop above
# opening the destination excel file
filename1 = "C:\\Users\\sebas\\OneDrive\\Software\\Python\\Excel_Auto_3\\Run_Folder\\master_test.xlsx"
wb2 = op.load_workbook(filename1)
ws2 = wb2.active
'''
'''
Has been included in the for loop above
# calculate total number of rows and
# columns in destination excel file
mr_dest = ws2.max_row
mc_dest = ws2.max_column

print(filename1)
print('mr_dest = ', mr_dest)
'''

'''
Has been included in the for loop above
# calculate total number of rows and
# columns in source excel file
mr = ws1.max_row
mc = ws1.max_column
'''

'''
Has been included in the for loop above
# copying the cell values from source
# excel file to destination excel file
for i in range (1, mr + 1):
    for j in range (1, mc + 1):
        # reading cell value from source excel file
        c = ws1.cell(row = i, column = j)
        print(c.value)
        # writing the read value to destination excel file, with
        # the change to have the values get added to the end of the
        # data already in the sheet
        ws2.cell(row = mr_dest + i, column = j).value = c.value

# saving the destination excel file
wb2.save(str(filename1))
'''
