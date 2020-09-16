# importing openpyxl module
import openpyxl as op
import os

acc_list = ['Bills', 'Emergency_+_Invest', 'Extra_Mortgage', 'Personal_Tax_&_SL', 'Savings', 'Trip', 'ACC_01', 'Credit_Card']
# An error occurs if there is no file with this file name in the working folder
# But the script still works


'''
# opening the source excel file
filename ="C:\\Users\\sebas\\OneDrive\\Software\\Python\\Excel Auto_2\\Transaction Files_Original\\Trans_files_working\\ACC_01.xlsx"
wb1 = op.load_workbook(filename)
ws1 = wb1.worksheets[0]
'''

'''
21/7/20 I have got this half working but only trip and savings is pulling through
and not all of the savings values are pulling through
'''

# opening the destination excel file
filename1 ="C:/Users/sebas/OneDrive/Software/Python/Excel_Auto_3/Run_Folder/master_test.xlsx"
print('destination filename =', filename1)
wb_master = op.load_workbook(filename1)
ws_master = wb_master.active
# calculate total number of rows
# in destination excel file
mr_master = ws_master.max_row
print('dest_max_row', mr_master)

wb1 = op.load_workbook('C:/Users/sebas/OneDrive/Software/Python/Excel_Auto_3/Correctly_Arranged_Sheets/ACC_01.xlsx')
ws1 = wb1.worksheets[0]

# calculate total number of rows and
# columns in source excel file
mr = ws1.max_row
mc = ws1.max_column

print("mr of first file = ", mr)
print("mc of first file = ", mc)


'''
# opening the source excel file
for (dirname, dirs, files) in os.walk('C:/Users/sebas/OneDrive/Software/Python/Excel_Auto_3/Correctly_Arranged_Sheets'):   #What is the fullstop here? Perhaps is means in the current directoy?
	for filename in files:
		wb1 = op.load_workbook(filename)
		ws1 = wb1.worksheets[0]

		# calculate total number of rows and
		# columns in source excel file
		mr = ws1.max_row
		mc = ws1.max_column

		print("mr = ", mr)
		print("mc = ", mc)
'''
'''
				# copying the cell values from source
				# excel file to destination excel file
				for i in range (1, mr + 1):
					for j in range (1, mc + 1):
						# reading cell value from source excel file
						c = ws1.cell(row = i, column = j)

						# writing the read value to destination excel file
						# This writes values after the last value in the destination sheet
						ws2.cell(row = mr2+i, column = j).value = c.value

				mr2 = ws2.max_row


				# saving the destination excel file
				wb2.save(str(filename1))
'''
