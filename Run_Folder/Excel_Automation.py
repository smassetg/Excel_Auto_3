#This program updates downloaded excel files, arranges them in the correct formate for the master sheet, combines them with the mastersheet and categorises items correctly

import pandas as pd
import openpyxl as op
import os

print('x=2')

'''
##################
Key Parameters for the script
#################
'''

nzhl_acc_list = [['38-9018-0879026-00','Emergency + Invest'],
['38-9018-0879026-01', 'Bills'],
['38-9018-0879026-02', 'Savings'],
['38-9018-0879026-03', 'Extra Mortgage'],
['38-9018-0879026-06', 'Trip'],
['38-9018-0879026-07', 'Personal Tax & SL']]

nzhl_acc_list_2 = [['Bills'], ['Emergency + Invest'], ['Extra Mortgage'], ['Personal Tax & SL'], ['Savings', 'Trip']]
nzhl_acc_list_3 = ['Bills', 'Emergency + Invest', 'Extra Mortgage', 'Personal Tax & SL', 'Savings', 'Trip']

'''
##################
Key Functions for the script
#################
'''

def west(filename):

	if 'ACC_01' in filename:
		account = 'ACC01'
	else:
		account = 'Credit_Card'

	wb = op.load_workbook(filename)
	sheet = wb.active
	mr = sheet.max_row
	# Arrange sheet in the format needed for the master sheet
	# Delete columns
	sheet.delete_cols(idx=5, amount=5)
	sheet.delete_cols(idx=1, amount=1)
	sheet.insert_cols(idx=2, amount=1)
	sheet.delete_rows(idx=1, amount=1) #delete the top row

	#This code updates all of the B column cells with the account name
	#by iterating through the rows and column position
	for x in range(1, mr):
		sheet.cell(column = 2, row = x, value = account)
		cell_num = 'C'+str(x) #Bizarre that I need to create this cell_num variable to use in the for loop
		sheet[cell_num].style = 'Currency' #This format of currency is bizarre - I could set my own but for now it does the trick inserting currency cells

	wb.save(filename)

	print('the function worked')

###############################
def nzhl(nzhl_account):

		wb = op.load_workbook(filename)
		sheet = wb.active
		mr = sheet.max_row
		move_range = 'G1:G'+str(mr) #This sets the move range and will update based on the number of max rows
		# Arrange sheet in the format needed for the master sheet
		# Delete columns
		sheet.delete_cols(17)
		sheet.delete_cols(5, 11)
		sheet.insert_cols(4, 2)
		sheet.move_range(move_range, cols=-2)
		sheet.delete_cols(1, 2)
		sheet.delete_rows(1)
		#This code updates all of the B column cells with the account name
		#by iterating through the rows and column position
		for x in range(1, mr):
			sheet.cell(column = 2, row = x, value = nzhl_account)
			cell_num = 'C'+str(x) #Bizarre that I need to create this cell_num variable to use in the for loop
			sheet[cell_num].style = 'Currency' #This format of currency is bizarre - I could set my own but for now it does the trick inserting currency cells
		wb.save(filename)

		print('the function worked for NZHL')
######################################

################
#Renames the files for boths Westpac and NZHL files
#These for statements updates the csv files into xlsx, updates the file name and removes the csv files from the folder

for (dirname, dirs, files) in os.walk('.'):   #What is the fullstop here? Perhaps is means in the current directoy?
	for filename in files:
		if filename.endswith('.csv'):
			if 'XXXX' in filename:
				account = 'Credit_Card'
			else:
				account = 'ACC_01'
			df = pd.read_csv(filename)
			df.to_excel(account + '.xlsx')
			os.remove(filename)

#These for statements updates the CSV files (NZHL) into xlsx, updates the file name and removes the CSV files from the folder
for (dirname, dirs, files) in os.walk('.'):   #What is the fullstop here? Perhaps is means in the current directoy?
	for filename in files:
		if filename.endswith('.CSV'):
			for rows in range(len(nzhl_acc_list)):
				if nzhl_acc_list[rows][0] in filename:
					nzhl_account = str(nzhl_acc_list[rows][1])
			df = pd.read_csv(filename)
			df.to_excel(nzhl_account + '_NZHL' + '.xlsx')
			os.remove(filename)


#These statements use functions to organise the sheets into the Master format
for (dirname, dirs, files) in os.walk('.'):   #What is the fullstop here? Perhaps is means in the current directoy?
	for filename in files:
		if filename.endswith('.xlsx'):
			print(filename)
			if 'NZHL' in filename:
				#nzhl_acc_names(filename)
				for items in range(len(nzhl_acc_list_3)):    #Please explain this. Without range(len...) I get the following error TypeError: list indices must be integers or slices, not list
				    if str(nzhl_acc_list_3[items]) in filename:
					    nzhl_account = str(nzhl_acc_list_3[items])   #Almost - there is one error with the mortgage condition
					    print(nzhl_account)
				nzhl(nzhl_account)
			else:
				west(filename) #bizarre - this identifies an error that filename ACC 01.xlsx does not exist - but there is no file name with that name? But the function seems to work


'''
################
#WB Arrangement WEST.py organises .csv files into the MASTER format
################

################
#WB Arrangement NZHL.py organises .CSV files into the MASTER format
################

#iteration3.py copies source data into the destination sheet

######################
#string_clean_NZHL.py removes the unwanted values within the spreadsheet - NZHL
######################

######################
#string_clean_WEST.py removes the unwanted values within the spreadsheet - NZHL
######################

To DO
 - Write code to copy each value into master sheet - this will be done through copying values into a 2D list and putting it into another sheet
 - Clean code so that it iterates through each excel file and excutes more cleanly - this will be done through the use of functions removing repetition


'''
